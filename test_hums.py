import pytest         as pt
import HUMS_objects   as ho
import Adams_object   as ao
import Eden_object    as eo
import file_functions as ff
import style_patch
import pdb #pdb.set_trace()

from pathlib  import Path
from openpyxl import load_workbook 

template_file  = "tested_files/1000691.036.AE ADAMS PVAI.xlsx"
tb1_file       = "tested_files/Step_1_lot_1727.csv"
tb2_file       = "tested_files/Step_2_lot_1727.csv"
accept_file    = "tested_files/HUMS_EDEN_LOT_01 25_20180201_145918.csv"
first_pn       = '1000636_test'
first_real_pn  = str(1000636)

# Objects init
data_files    = ho.Datafiles(template_file, tb1_file, tb2_file, accept_file)
batch         = ho.Batch()
batch.get_products(data_files.files())
batch.product_type_id()

#defining batch type
if batch.product_type == 'E':
    product = eo.Eden('E00217', data_files.files())
elif batch.product_type == 'A':
    product = ao.Adams('A00155', data_files.files())

# Running functions using files
batch.generated_pn(first_real_pn)
product.generate_pv(template_file, first_pn, batch.products)
product.get_attributs_from_acceptance(data_files.files())
product.get_consumption(data_files.files())

# Workbook management
wb = {}
wb[product.SN]    = load_workbook(filename= product.pv)
ws                = wb[product.SN].active
ff.img_import(wb, product.SN)

# Running functions using workbooks
product.batch_fill_pv(batch.products, ws)
product.pv_header(first_real_pn, ws)

# Datafile tests
def test_datafiles():
    assert len(data_files.files()) == 4
def test_datafiles1():
    assert data_files.files()[0] == template_file
def test_file_control():
    assert data_files.file_control() == [['E00215','E00227','E00236'],{'E00231':'seen once', 'E00232':'not seen', 'E00234':'seen 3 times', 'E00236':'not seen'}, ['E00232']]

# Hums batch test
def test_get_products():
	assert len(batch.products) == 24
def test_get_products1():
	assert batch.products[0]   == 'E00213'
def test_generated_pn():
    assert len(batch.pn)       == len(batch.products)
def test_generated_pn1():
    assert batch.pn[0]         == '1000636'
    assert batch.pn[10]        == '1000646'
def test_product_type():
	assert batch.product_type  == 'E'


# Hums tests
def test_get_attributs():
	assert product.hums_attributs['Resultat de la mesure de vibration sur X'] == 2.17
def test_get_attributs1():	
	assert product.hums_attributs['Duree maximale des chocs enregistrés']     == 22.549768

def test_consumption_sleep():
    assert product.hums_attributs['conso_sleep'] == 235.3
def test_consumption_acq():
    assert product.hums_attributs['conso_acq']   == 4420.8
def test_consumption_stock():
    assert product.hums_attributs['conso_stock'] == 15.5

def test_batch_fill_pv():
	assert ws['D13'].value == 'E00213' 

def test_pv_header():
	assert ws['D1'].value  == 'PROCES VERVAL D’ACCEPTATION\nN°' + first_real_pn

def test_generate():
    file = Path("1000636_test.AA_E00217_PVAI.xlsx") 
    assert file.is_file()

if __name__=='__main__':
    pt.main()
