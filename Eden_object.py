import file_functions  as ff
import HUMS_objects    as ho
import functools       as ft
import pdb #pdb.set_trace()
import logging

from operator import itemgetter
from openpyxl import load_workbook

log = logging.getLogger(__name__)

class Eden(ho.Hums):
    def __init__(self, SN, files):
        ho.Hums.__init__(self, SN, files)
        
        #defining test results of eden
        #4 types of tets:  S = status, T = thresholds, R = tolerence ref, P = % ref
        self.waited_test_result['T_conso_veil']  = [215, 250]
        self.waited_test_result['T_conso_perio'] = [4250, 4500]
        self.waited_test_result['T_conso_stock'] = [10, 20]
        self.waited_test_result['T_compr_joint'] = [73.2, 74.2]
        self.waited_test_result['T_poids']       = [0, 1250]
        self.waited_test_result['T_conti_cable'] = [0, 25]
        self.waited_test_result['T_conti_capot'] = [0, 5]
        self.waited_test_result['T_long_cable']  = [340, 360]
        self.waited_test_result['R_pressure']    = 50
        self.waited_test_result['R_temperature'] = 1
        self.waited_test_result['R_humidity']    = 5
        self.waited_test_result['P_vibration']   = 10
        self.waited_test_result['R_choc']        = 1
        self.waited_test_result['P_choc']        = 10
        self.waited_test_result['P_choc_trans']  = [0, 20]

    def fill_pv(self, ws):
        wb         = {}
        wb[self.SN]= load_workbook(filename= self.pv)
        ws         = wb[self.SN].active
        ff.img_import(wb, self.SN)

        self.fill_eden(ws)
        self.test_eden(ws)

        wb[self.SN].save(filename = self.pv)


    def fill_eden(self, ws):
        #Conso and other tests
        self.writing_tester(ws, 'F108', 'conso_sleep', ' µA')
        self.writing_tester(ws, 'F109', 'conso_acq', ' µA')
        self.writing_tester(ws, 'F110', 'conso_stock', ' µA')
        self.writing_tester(ws, 'F112', 'Compression joints piles', ' mm')
        self.writing_tester(ws, 'F163', 'Poids', 'g')
        self.writing_tester(ws, 'F167', 'Continuite electrique capot', ' mΩ')
        self.writing_tester(ws, 'F166', 'Continuite electrique cable', ' mΩ')
        self.writing_tester(ws, 'F159', 'Longueur de cable', ' mm')
    

        #Temperature
        self.writing_tester(ws, 'F172', 'Temperature Hums 1', '°C', 1)
        self.writing_tester(ws, 'F173', 'Temperature Hums 3', '°C', 1)
        self.writing_tester(ws, 'F174', 'Temperature Hums 2', '°C', 1)

        self.writing_tester(ws, 'D172', 'Temperature Ref 1', '°C', 1)
        self.writing_tester(ws, 'D173', 'Temperature Ref 3', '°C', 1)
        self.writing_tester(ws, 'D174', 'Temperature Ref 2', '°C', 1)

        #Humidity
        self.writing_tester(ws, 'F176', 'Humidite Hums 1', '%', 1)
        self.writing_tester(ws, 'F177', 'Humidite Hums 2', '%', 1)
        self.writing_tester(ws, 'F178', 'Humidite Hums 4', '%', 1)

        self.writing_tester(ws, 'D176', 'Humidite Ref 1', '%', 1)
        self.writing_tester(ws, 'D177', 'Humidite Ref 2', '%', 1)
        self.writing_tester(ws, 'D178', 'Humidite Ref 4', '%', 1)

        #Pressure
        self.writing_tester(ws, 'F180', 'Pression Hums 1', ' hPa', 1)
        self.writing_tester(ws, 'F181', 'Pression Hums 3', ' hPa', 1)
        self.writing_tester(ws, 'F182', 'Pression Hums 2', ' hPa', 1)

        self.writing_tester(ws, 'D180', 'Pression Ref 1', ' hPa', 1)
        self.writing_tester(ws, 'D181', 'Pression Ref 3', ' hPa', 1)
        self.writing_tester(ws, 'D182', 'Pression Ref 2', ' hPa', 1)

        #Vibrations
        self.writing_tester(ws, 'F118', 'Resultat de la mesure de vibration sur X', ' grms', 2)
        self.writing_tester(ws, 'F119', 'Resultat de la mesure de vibration sur Y', ' grms', 2)
        self.writing_tester(ws, 'F120', 'Resultat de la mesure de vibration sur Z', ' grms', 2)

        self.writing_tester(ws, 'D118', 'Resultat de la valeur de reference de vibration sur X', ' grms', 2)
        self.writing_tester(ws, 'D119', 'Resultat de la valeur de reference de vibration sur Y', ' grms', 2)
        self.writing_tester(ws, 'D120', 'Resultat de la valeur de reference de vibration sur Z', ' grms', 2)

        #Chocs
        self.writing_tester(ws, 'F189', 'Valeur mesuree des chocs sur l\'axe X', ' g', 2)
        self.writing_tester(ws, 'D189', 'Valeur de reference des chocs sur l\'axe X', ' g', 2)
        self.writing_tester(ws, 'F190', 'Shock (transverse X) axe Y', ' g', 2)
        self.writing_tester(ws, 'F191', 'Shock (transverse X) axe Z', ' g', 2)
        self.writing_tester(ws, 'F195', 'Duree choc axe X (ms)', ' ms', 1)

        self.writing_tester(ws, 'F196', 'Valeur mesuree des chocs sur l\'axe Y', ' g', 2)
        self.writing_tester(ws, 'D196', 'Valeur de reference des chocs sur l\'axe Y', ' g', 2)
        self.writing_tester(ws, 'F197', 'Shock (transverse Y) axe X', ' g', 2)
        self.writing_tester(ws, 'F198', 'Shock (transverse Y) axe Z', ' g', 2)
        self.writing_tester(ws, 'F199', 'Duree choc axe Y (ms)', ' ms', 1)

        self.writing_tester(ws, 'F200', 'Valeur mesuree des chocs sur l\'axe Z', ' g', 2)
        self.writing_tester(ws, 'D200', 'Valeur de reference des chocs sur l\'axe Z', ' g', 2)
        self.writing_tester(ws, 'F201', 'Shock (transverse Z) axe X', ' g', 2)
        self.writing_tester(ws, 'F202', 'Shock (transverse Z) axe Y', ' g', 2)
        self.writing_tester(ws, 'F203', 'Duree choc axe Z (ms)', ' ms', 1)


    def test_eden(self, ws):
        #Thresholded values checked
        self.threshold_check(ws, 'T_conso_veil' , 'conso_sleep', 'G108')
        self.threshold_check(ws, 'T_conso_perio', 'conso_acq'  , 'G109')
        self.threshold_check(ws, 'T_conso_stock', 'conso_stock', 'G110')
        self.threshold_check(ws, 'T_compr_joint', 'Compression joints piles', 'G112')
        self.threshold_check(ws, 'T_poids', 'Poids', 'G163')
        self.threshold_check(ws, 'T_conti_capot', 'Continuite electrique capot', 'G167')
        self.threshold_check(ws, 'T_conti_cable', 'Continuite electrique cable', 'G166')
        self.threshold_check(ws, 'T_long_cable', 'Longueur de cable', 'G159')
        #Tolerenced values checked
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 1', 'Temperature Ref 1', 'G172')
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 3', 'Temperature Ref 3', 'G173')
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 2', 'Temperature Ref 2', 'G174')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 1', 'Humidite Ref 1', 'G176')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 2', 'Humidite Ref 2', 'G177')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 4', 'Humidite Ref 4', 'G178')
        self.tolerence_check(ws, 'R_pressure', 'Pression Hums 1', 'Pression Ref 1', 'G180')
        self.tolerence_check(ws, 'R_pressure', 'Pression Hums 3', 'Pression Ref 3', 'G181')
        self.tolerence_check(ws, 'R_pressure', 'Pression Hums 2', 'Pression Ref 2', 'G182')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur X', 'Resultat de la valeur de reference de vibration sur X', 'G118', 'PERCENT')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur Y', 'Resultat de la valeur de reference de vibration sur Y', 'G119', 'PERCENT')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur Z', 'Resultat de la valeur de reference de vibration sur Z', 'G120', 'PERCENT')
        
        self.tolerence_check(ws, 'R_choc', 'Valeur mesuree des chocs sur l\'axe X', 'Valeur de reference des chocs sur l\'axe X', 'G189', 'BOTH', 'P_choc')
        self.tolerence_check(ws, 'R_choc', 'Valeur mesuree des chocs sur l\'axe Y', 'Valeur de reference des chocs sur l\'axe Y', 'G196', 'BOTH', 'P_choc')
        self.tolerence_check(ws, 'R_choc', 'Valeur mesuree des chocs sur l\'axe Z', 'Valeur de reference des chocs sur l\'axe Z', 'G200', 'BOTH', 'P_choc')
        
        try:
            #pdb.set_trace()
            perc_trans_yx = 100*self.hums_attributs['Shock (transverse X) axe Y']/self.hums_attributs['Shock (transverse X) axe X'] if self.hums_attributs['Shock (transverse X) axe Y'] else 0
            perc_trans_zx = 100*self.hums_attributs['Shock (transverse X) axe Z']/self.hums_attributs['Shock (transverse X) axe X'] if self.hums_attributs['Shock (transverse X) axe Z'] else 0
            ws['G190'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_yx) else 'NOK'
            ws['G191'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_zx) else 'NOK'
        except Exception:
            log.error('Transverse X result failed on: {}'.format(self.SN))

        try:
            perc_trans_xy = 100*self.hums_attributs['Shock (transverse Y) axe X']/self.hums_attributs['Shock (transverse Y) axe Y'] if self.hums_attributs['Shock (transverse Y) axe X'] else 0
            perc_trans_zy = 100*self.hums_attributs['Shock (transverse Y) axe Z']/self.hums_attributs['Shock (transverse Y) axe Y'] if self.hums_attributs['Shock (transverse Y) axe Z'] else 0
            ws['G197'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_xy) else 'NOK'
            ws['G198'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_zy) else 'NOK'
        except Exception:
            log.error('Transverse Y result failed on: {}'.format(self.SN))

        try:
            perc_trans_xz = 100*self.hums_attributs['Shock (transverse Z) axe X']/self.hums_attributs['Shock (transverse Z) axe Z'] if self.hums_attributs['Shock (transverse Z) axe X'] else 0
            perc_trans_yz = 100*self.hums_attributs['Shock (transverse Z) axe Y']/self.hums_attributs['Shock (transverse Z) axe Z'] if self.hums_attributs['Shock (transverse Z) axe Y'] else 0
            ws['G201'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_xz) else 'NOK'
            ws['G202'] = 'OK' if self.threshold_check(ws, 'P_choc_trans', perc_trans_yz) else 'NOK'
        except Exception:
            log.error('Transverse Z result failed on: {}'.format(self.SN))

            
            

        

        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse X) axe Y', 'Shock (transverse X) axe X', 'G190', 'PERCENT')
        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse X) axe Z', 'Shock (transverse X) axe X', 'G191', 'PERCENT')
        
        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse Y) axe X', 'Shock (transverse Y) axe Y', 'G197', 'PERCENT')
        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse Y) axe Z', 'Shock (transverse Y) axe Y', 'G198', 'PERCENT')
        
        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse Z) axe X', 'Shock (transverse Z) axe Z', 'G201', 'PERCENT')
        #self.tolerence_check(ws, 'P_choc_trans', 'Shock (transverse Z) axe Y', 'Shock (transverse Z) axe Z', 'G202', 'PERCENT')












