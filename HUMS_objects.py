import pyexcel        as pe
import pandas         as pd
import table_utils    as tu
import file_functions as ff
import style_patch
import pdb

from openpyxl import load_workbook
from numpy import mean

log = tu.logging_manager()

class Datafiles:
    def __init__(self, file0, file1, file2, file3):
        self.file0 = file0 # Template
        self.file1 = file1 # TB1
        self.file2 = file2 # TB2
        self.file3 = file3 # Acceptance

    def files(self):
        return [self.file0, self.file1, self.file2, self.file3]

    def file_control(self):
        tb1_list_double  = ff.tb1_check(self.file1)
        tb2_dict_error   = ff.tb2_check(self.file1, self.file2)
        acc_list_missing = ff.acceptance_check(self.file1, self.file3)
        #logging file problems
        if tb1_list_double:
            log.error('Products {} appears twice or more in {}, this file should contain only uniq product result'.format(tb1_list_double, self.file1))
        if tb2_dict_error:
            log.warning('The folowing problems: {} appears in {}, this file should contain each product twice'.format(tb2_dict_error, self.file2))
        if acc_list_missing:
            log.warning('The folowing products are missing: {} in {}, this file should contain the same products than the 1rs testbench step'.format(acc_list_missing, self.file3))

        return [tb1_list_double, tb2_dict_error, acc_list_missing]

class Batch:
    def __init__(self):
        self.products     = []
        self.pn           = []
        self.product_type = ''

    def get_products(self, files):
        rf = pd.read_csv(files[1])
        self.products = list(rf['SN'])

    def generated_pn(self, first_pn):
        self.pn = ff.generate_pn(first_pn, self.products)

    def product_type_id(self):
    	batch = set()
    	for p in self.products:
    			batch.add(str(p[0]))
    	if len(batch) == 1:
    		self.product_type = str(p[0])
    	else:
    		self.product_type = 'inconsistante batch'
    		log.error('The batch should contain only one type of products (EDEN or ADAMS), actual batch: {}'.format(self.products))		

class Hums:
    def __init__(self, SN, files):
        self.SN                  = SN
        self.hums_attributs      = {}
        self.test_bench_1_row    = tu.csv_row_search(self.SN, 'SN', files[1], ',')
        self.test_bench_2_row    = tu.csv_row_search(self.SN, 'SN', files[2], ',')
        self.product_row         = tu.csv_row_search(self.SN, 'Numero de serie', files[3], ';')
        self.pv                  = ''
        self.waited_test_result  = {}

    def get_attributs_from_acceptance(self, files):
        data_dict = pe.get_dict(file_name= files[3], delimiter=';')
        for key, value in data_dict.items():
            self.hums_attributs[key] = value[self.product_row]
   
    def generate_pv(self, template, PN, batch):
        wb          = {}
        wb[self.SN] = load_workbook(filename= template)
        ws          = wb[self.SN].active
        ff.img_import(wb, self.SN)
        self.pv     =  PN  + '.AA_' + self.SN + '_PVAI.xlsx'
        self.batch_fill_pv(batch, ws)
        self.pv_header(PN, ws)

        wb[self.SN].save(filename = self.pv) 

    def batch_fill_pv(self, batch, worksheet):
        pv_batch_columns = ['D','E','F','G', 'H']
        pv_batch_lines   = range(13, 17)
        
        worksheet['D11'] = len(batch)
        for ind_col,let in enumerate(pv_batch_columns):
            for ind_line,num in enumerate(pv_batch_lines):
                if (ind_col*4 + ind_line) < len(batch):
                    worksheet[let + str(num)] = batch[ind_col*4 + ind_line]
    
    def pv_header(self, pn, worksheet):
        worksheet['D1']  = 'PROCES VERVAL D’ACCEPTATION\nN°' + pn
        worksheet['D31'] = 'PROCES VERVAL D’ACCEPTATION INDIVIDUEL\nN°' + pn
        worksheet['D38'] = 'NUMERO DE SERIE : ' + self.SN
    
    def writing_tester(self, worksheet, cell_value, attribut, unit='', rounding=0):
            try :
                if self.hums_attributs[attribut]:
                    if rounding:
                        worksheet[cell_value] = str(round(self.hums_attributs[attribut], rounding)) + unit
                    else:
                        worksheet[cell_value] = str(self.hums_attributs[attribut]) + unit
                else:
                    log.warning('No value found for:, {} on product : {}'.format(attribut, self.SN))
            except Exception:
                log.error('Error during calculation of:, {} on product : {}'.format(attribut, self.SN))
                    
    def threshold_check(self, ws, threshold, tested_attribut, result_cell):
        try:
            if self.hums_attributs[tested_attribut]:
                value = round(self.hums_attributs[tested_attribut], 2)
            else:
                ws[result_cell] = 'NOK'
                log.warning('Testing NOK for, {} on product {}'.format(threshold, self.SN))
                return

            if self.waited_test_result[threshold][0] <= value <= self.waited_test_result[threshold][1]:
                ws[result_cell] = 'OK'
            else:
                ws[result_cell] = 'NOK'
                log.warning('Testing NOK for, {} on product {}'.format(threshold, self.SN))
        except Exception:
            log.error('Testing value failled for, {} on product {}'.format(threshold, self.SN))

    def status_check(self, ws, status, result_cell):
        try:
            out_cell = result_cell
            inp_cell = result_cell.replace('G', 'F')
            if ws[inp_cell].value == self.waited_test_result[status]:
                ws[out_cell] = 'OK'
            else:
                ws[out_cell] = 'NOK'
                log.warning('Testing NOK for, {} on product {}'.format(threshold, self.SN))
        except Exception:
            log.error('Status value failled for cell, {} on product {}'.format(result_cell, self.SN))

    def tolerence_check(self, ws, tolerence, tested_attribut, reference, result_cell, mode='VALUE', tolerence1=0):
        try:
            if self.hums_attributs[tested_attribut]:
                value = round(self.hums_attributs[tested_attribut], 2)
            else:
                ws[result_cell] = 'NOK'
                log.warning('Testing NOK for, {} on product {}'.format(tolerence, self.SN))
                return

            # tolerence is defined on the mode of the test
            # PERCENT mode calculates the treshold using the tolerence as a percent of reference value
            # VALUE mode calculates the treshold as a normale treshold : + and - the treshold value defined
            # BOTH mode calculates the treshold by adding both modes
            if mode == 'PERCENT':
                tol = (self.waited_test_result[tolerence]/100)*abs(self.hums_attributs[reference])
            elif mode == 'VALUE':
                tol = self.waited_test_result[tolerence]
            elif mode == 'BOTH':
            	tol = (self.waited_test_result[tolerence1]/100)*abs(self.hums_attributs[reference]) + self.waited_test_result[tolerence]

            if abs(self.hums_attributs[reference]) - tol <= abs(value) <=  abs(self.hums_attributs[reference]) + tol:
                ws[result_cell] = 'OK'
            else :
                ws[result_cell] = 'NOK'
                log.warning('Testing NOK for, {} on product {}'.format(tolerence, self.SN))

        except Exception:
            log.error('Testing value failled for, {} on product {}'.format(tolerence, self.SN))

    def get_consumption(self, file):
        #value filtering tresholds
        acq_low_fltr    = 0.002500
        acq_high_fltr   = 0.006000
        sleep_low_fltr  = 0
        sleep_high_fltr = 0.000500
        row             = self.test_bench_1_row

        #acquisition mode and sleep mode
        start_col   = tu.csv_col_search('Consumption 1' , file[1], ',')
        end_col     = tu.csv_col_search('Consumption 40', file[1], ',')
        total_conso = tu.rounding_list(tu.get_list_from_csv_row(file[1], row, start_col, end_col), 10)
        acqui_mode  = tu.filter_conso(total_conso, acq_low_fltr, acq_high_fltr, 'acqui')
        sleep_mode  = tu.filter_conso(total_conso, sleep_low_fltr, sleep_high_fltr, 'sleep')
        
        #stock mode
        stock_start_col = tu.csv_col_search('Consumption Ds1' , file[1], ',')
        stock_end_col   = tu.csv_col_search('Consumption Ds11' , file[1], ',')
        stock_mode      = tu.rounding_list(tu.get_list_from_csv_row(file[1], row, stock_start_col, stock_end_col), 10)
        
        #Rounding values
        conso_sleep = round(mean(sleep_mode)*10**6, 1)
        conso_acq   = round(mean(acqui_mode)*10**6, 1)
        conso_stock = round(mean(stock_mode)*10**6, 1)

        #controling data & saving
        if not sleep_mode or not acqui_mode:
            log.error('Consumption error:, sleep mode is {} and acquisition mode is {} on product {}'.format(conso_sleep, conso_acq, self.SN))
            conso_sleep = 0
            conso_acq   = 0
        
        if len(set(sleep_mode + acqui_mode) - set(total_conso)) > 3:
            log.error('Consumption error:, too many kicked values on product {}'.format(self.SN))
            conso_sleep = 0
            conso_acq   = 0
                     
        self.hums_attributs['conso_sleep'] = conso_sleep
        self.hums_attributs['conso_acq']   = conso_acq
        self.hums_attributs['conso_stock'] = conso_stock
            
    





