import file_functions  as ff
import HUMS_objects    as ho
import pdb #pdb.set_trace()
import logging

from operator import itemgetter
from openpyxl import load_workbook

log = logging.getLogger(__name__)

class Adams(ho.Hums):
    def __init__(self, SN, files):
        ho.Hums.__init__(self, SN, files)
        
        #defining test results of adams
        #4 types of tets:  S = status, T = thresholds, R = tolerence ref, P = % ref
        self.waited_test_result['T_conso_veil']  = [225, 275]
        self.waited_test_result['T_conso_perio'] = [4000, 4500]
        self.waited_test_result['T_conso_stock'] = [10, 30]
        self.waited_test_result['T_compr_joint'] = [3.15, 4.15]
        self.waited_test_result['T_poids']       = [0, 1250]
        self.waited_test_result['S_gabarit_dim'] = 'OK'
        self.waited_test_result['T_conti_libre'] = [0, 1]
        self.waited_test_result['T_conti_monte'] = [0, 1.5]
        self.waited_test_result['S_led']         = 'OK'
        self.waited_test_result['R_temperature'] = 1
        self.waited_test_result['R_humidity']    = 5
        self.waited_test_result['P_vibration']   = 10
        self.waited_test_result['T_src']         = [0, 10]
        self.waited_test_result['T_src_trans']   = [0, 15]
        self.waited_test_result['P_src_vect']    = 10

    def fill_pv(self, ws):
        wb         = {}
        wb[self.SN]= load_workbook(filename= self.pv)
        ws         = wb[self.SN].active
        ff.img_import(wb, self.SN)

        self.fill_adams(ws)
        self.test_adams(ws)

        wb[self.SN].save(filename = self.pv)


    def fill_adams(self, ws):
        #Conso and other tests
        self.writing_tester(ws, 'F107', 'conso_sleep', ' µA')
        self.writing_tester(ws, 'F108', 'conso_acq', ' µA')
        self.writing_tester(ws, 'F109', 'conso_stock', ' µA')
        self.writing_tester(ws, 'F111', 'Compression joints piles', ' mm')
        self.writing_tester(ws, 'F137', 'Poids', 'g')
        self.writing_tester(ws, 'F139', 'Resultat du test gabarit')
        self.writing_tester(ws, 'F140', 'Resultat de la valeur de mesure du HUMS libre au milliohmetre', ' mΩ')
        self.writing_tester(ws, 'F141', 'Resultat de la valeur de mesure du HUMS monte au milliohmetre', ' mΩ')

        #Led button
        if self.hums_attributs['Resultat test bouton vert'] == 'OK' and self.hums_attributs['Resultat test bouton rouge'] == 'OK':
            self.writing_tester(ws, 'F152', 'Resultat test bouton vert')
        else:
            ws['F152'] = 'NOK'        

        #Temperature
        self.writing_tester(ws, 'F154', 'Temperature Hums 2', '°C', 1)
        self.writing_tester(ws, 'F155', 'Temperature Hums 3', '°C', 1)
        self.writing_tester(ws, 'F156', 'Temperature Hums 1', '°C', 1)

        self.writing_tester(ws, 'D154', 'Temperature Ref 2', '°C', 1)
        self.writing_tester(ws, 'D155', 'Temperature Ref 3', '°C', 1)
        self.writing_tester(ws, 'D156', 'Temperature Ref 1', '°C', 1)

        #Humidity
        self.writing_tester(ws, 'F158', 'Humidite Hums 2', '%', 1)
        self.writing_tester(ws, 'F159', 'Humidite Hums 1', '%', 1)
        self.writing_tester(ws, 'F160', 'Humidite Hums 3', '%', 1)

        self.writing_tester(ws, 'D158', 'Humidite Ref 2', '%', 1)
        self.writing_tester(ws, 'D159', 'Humidite Ref 1', '%', 1)
        self.writing_tester(ws, 'D160', 'Humidite Ref 3', '%', 1)

        #Vibrations
        self.writing_tester(ws, 'F166', 'Resultat de la mesure de vibration sur X', ' grms', 2)
        self.writing_tester(ws, 'F167', 'Resultat de la mesure de vibration sur Y', ' grms', 2)
        self.writing_tester(ws, 'F168', 'Resultat de la mesure de vibration sur Z', ' grms', 2)

        self.writing_tester(ws, 'D166', 'Resultat de la valeur de reference de vibration sur X', ' grms', 2)
        self.writing_tester(ws, 'D167', 'Resultat de la valeur de reference de vibration sur Y', ' grms', 2)
        self.writing_tester(ws, 'D168', 'Resultat de la valeur de reference de vibration sur Z', ' grms', 2)

        #Chocs
        self.writing_tester(ws, 'F170', 'Ecart max de l\'analyse SRC X', '%', 2)
        self.writing_tester(ws, 'F172', 'Ecart max de l\'analyse SRC Y', '%', 2)
        self.writing_tester(ws, 'F174', 'Ecart max de l\'analyse SRC Z', '%', 2)

        ws['D170'] = '-'
        self.writing_tester(ws, 'D171', 'Valeur de reference des chocs post-calibration sur l\'axe X', 'g', 2)
        ws['D172'] = '-'
        self.writing_tester(ws, 'D173', 'Valeur de reference des chocs post-calibration sur l\'axe Y', 'g', 2)
        ws['D174'] = '-'
        self.writing_tester(ws, 'D175', 'Valeur de reference des chocs post-calibration sur l\'axe Z', 'g', 2)

        percent_trans    = lambda trans, choc: 100*abs(trans/choc)
        percent_choc     = lambda value, ref : 100*abs((abs(value)-abs(ref))/ref)
        
        try:
            max_transverse_x = max(percent_trans(self.hums_attributs['Shock (transverse X) axe Y'], self.hums_attributs['Shock (transverse X) axe X']) , percent_trans(self.hums_attributs['Shock (transverse X) axe Z'], self.hums_attributs['Shock (transverse X) axe X']))
            choc_vector_x    = (self.hums_attributs['Shock (transverse X) axe X']**2 + self.hums_attributs['Shock (transverse X) axe Y']**2 +self.hums_attributs['Shock (transverse X) axe Z']**2)**(1/2)
            perc_vector_x    = percent_choc(choc_vector_x, self.hums_attributs['Valeur de reference des chocs post-calibration sur l\'axe X'])
            ws['F171']       = str(round(max_transverse_x, 1)) + '%' + '\n'+ str(round(perc_vector_x, 1)) + '%'

            max_transverse_y = max(percent_trans(self.hums_attributs['Shock (transverse Y) axe X'], self.hums_attributs['Shock (transverse Y) axe Y']),percent_trans(self.hums_attributs['Shock (transverse Y) axe Z'], self.hums_attributs['Shock (transverse Y) axe Y']))
            choc_vector_y    = (self.hums_attributs['Shock (transverse Y) axe X']**2 + self.hums_attributs['Shock (transverse Y) axe Y']**2 +self.hums_attributs['Shock (transverse Y) axe Z']**2)**(1/2)
            perc_vector_y    = percent_choc(choc_vector_y, self.hums_attributs['Valeur de reference des chocs post-calibration sur l\'axe Y'])
            ws['F173']       = str(round(max_transverse_y, 1)) + '%' + '\n'+ str(round(perc_vector_y, 1)) + '%'

            max_transverse_z = max(percent_trans(self.hums_attributs['Shock (transverse Z) axe X'], self.hums_attributs['Shock (transverse Z) axe Z']), percent_trans(self.hums_attributs['Shock (transverse Z) axe Y'], self.hums_attributs['Shock (transverse Z) axe Z']))
            choc_vector_z    = (self.hums_attributs['Shock (transverse Z) axe X']**2 + self.hums_attributs['Shock (transverse Z) axe Y']**2 +self.hums_attributs['Shock (transverse Z) axe Z']**2)**(1/2)
            perc_vector_z    = percent_choc(choc_vector_z, self.hums_attributs['Valeur de reference des chocs post-calibration sur l\'axe Z'])
            ws['F175']       = str(round(max_transverse_z, 1)) + '%' + '\n'+ str(round(perc_vector_z, 1)) + '%'

        except Exception:
            max_transverse_x = 0
            log.error('Max transverse calculation error on product: {}'.format(self.SN))

        
        #Ecretage
        try:
            ecret_min_low = min(self.hums_attributs['Ecretage -20 axe X'], self.hums_attributs['Ecretage -20 axe Y'], self.hums_attributs['Ecretage -20 axe Z'])
            ecret_min_hig = min(self.hums_attributs['Ecretage 70 axe X'] , self.hums_attributs['Ecretage 70 axe Y'] , self.hums_attributs['Ecretage 70 axe Z'] )
            ws['D215'] = str(round(ecret_min_low, 1)) + ' g'
            ws['D216'] = str(round(ecret_min_hig, 1)) + ' g'
        except Exception:
            ecret_min_low = 0
            ecret_min_hig = 0
            log.error('Ecretage minimal calculation error on product: {}'.format(self.SN))
        
        #Retrieving axis of mininimum ecretage
        if ecret_min_low and ecret_min_hig:
            dic_ecrt_low  = {key:self.hums_attributs[key] for key in ['Ecretage -20 axe X', 'Ecretage -20 axe Y', 'Ecretage -20 axe Z']}
            dic_ecrt_high = {key:self.hums_attributs[key] for key in ['Ecretage 70 axe X' , 'Ecretage 70 axe Y' , 'Ecretage 70 axe Z'] }
            ws['F215']    = str(min(dic_ecrt_low.items(),  key=itemgetter(1))[0][13:])
            ws['F216']    = str(min(dic_ecrt_high.items(), key=itemgetter(1))[0][12:])

    def test_adams(self, ws):
        #Thresholded values checked
        self.threshold_check(ws, 'T_conso_veil' , 'conso_sleep', 'G107')
        self.threshold_check(ws, 'T_conso_perio', 'conso_acq'  , 'G108')
        self.threshold_check(ws, 'T_conso_stock', 'conso_stock', 'G109')
        self.threshold_check(ws, 'T_compr_joint', 'Compression joints piles', 'G111')
        self.threshold_check(ws, 'T_poids', 'Poids', 'G137')
        self.threshold_check(ws, 'T_conti_libre', 'Resultat de la valeur de mesure du HUMS libre au milliohmetre', 'G140')
        self.threshold_check(ws, 'T_conti_monte', 'Resultat de la valeur de mesure du HUMS monte au milliohmetre', 'G141')
        #Tolerenced values checked
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 2', 'Temperature Ref 2', 'G154')
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 3', 'Temperature Ref 3', 'G155')
        self.tolerence_check(ws, 'R_temperature', 'Temperature Hums 1', 'Temperature Ref 1', 'G156')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 2', 'Humidite Ref 2', 'G158')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 1', 'Humidite Ref 1', 'G159')
        self.tolerence_check(ws, 'R_humidity', 'Humidite Hums 3', 'Humidite Ref 3', 'G160')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur X', 'Resultat de la valeur de reference de vibration sur X', 'G166', 'PERCENT')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur Y', 'Resultat de la valeur de reference de vibration sur Y', 'G167', 'PERCENT')
        self.tolerence_check(ws, 'P_vibration', 'Resultat de la mesure de vibration sur Z', 'Resultat de la valeur de reference de vibration sur Z', 'G168', 'PERCENT')
        self.threshold_check(ws, 'T_src', 'Ecart max de l\'analyse SRC X', 'G170')
        self.threshold_check(ws, 'T_src', 'Ecart max de l\'analyse SRC Y', 'G172')
        self.threshold_check(ws, 'T_src', 'Ecart max de l\'analyse SRC Z', 'G174')
        self.status_check(ws, 'S_gabarit_dim', 'G139')
        self.status_check(ws, 'S_led', 'G152')

        # Max of transverse axis identification
        try:
            #pdb.set_trace()
            dic_transverse_x = {key:self.hums_attributs[key] for key in ['Shock (transverse X) axe Y', 'Shock (transverse X) axe Z']}
            dic_transverse_y = {key:self.hums_attributs[key] for key in ['Shock (transverse Y) axe X', 'Shock (transverse Y) axe Z']}
            dic_transverse_z = {key:self.hums_attributs[key] for key in ['Shock (transverse Z) axe X', 'Shock (transverse Z) axe Y']}
            col_name_x   = str(max(dic_transverse_x.items(), key=itemgetter(1))[0])
            col_name_y   = str(max(dic_transverse_y.items(), key=itemgetter(1))[0])
            col_name_z   = str(max(dic_transverse_z.items(), key=itemgetter(1))[0])
            perc_trans_x = 100*self.hums_attributs[col_name_x]/self.hums_attributs['Shock (transverse X) axe X']
            perc_trans_y = 100*self.hums_attributs[col_name_y]/self.hums_attributs['Shock (transverse Y) axe Y']
            perc_trans_z = 100*self.hums_attributs[col_name_z]/self.hums_attributs['Shock (transverse Z) axe Z']
            

            if self.threshold_check(ws, 'T_src_trans', perc_trans_x) or self.tolerence_check(ws, 'P_src_vect', [self.hums_attributs['Shock (transverse X) axe X'], self.hums_attributs['Shock (transverse X) axe Y'], self.hums_attributs['Shock (transverse X) axe Z']], self.hums_attributs['Valeur de reference des chocs post-calibration sur l\'axe X']):
                ws['G171'] = 'OK'
            else :
                ws['G171'] = 'NOK'
                log.warning('Testing NOK for transverse x on product {}'.format(self.SN))

            if self.threshold_check(ws, 'T_src_trans', perc_trans_y) or self.tolerence_check(ws, 'P_src_vect', [self.hums_attributs['Shock (transverse Y) axe X'], self.hums_attributs['Shock (transverse Y) axe Y'], self.hums_attributs['Shock (transverse Y) axe Z']], self.hums_attributs['Valeur de reference des chocs post-calibration sur l\'axe Y']):
                ws['G173'] = 'OK'
            else :
                ws['G173'] = 'NOK'
                log.warning('Testing NOK for transverse y on product {}'.format(self.SN))

            if self.threshold_check(ws, 'T_src_trans', perc_trans_z) or self.tolerence_check(ws, 'P_src_vect', [self.hums_attributs['Shock (transverse Z) axe X'], self.hums_attributs['Shock (transverse Z) axe Y'], self.hums_attributs['Shock (transverse Z) axe Z']], 'Valeur de reference des chocs post-calibration sur l\'axe Z'):
                ws['G175'] = 'OK'
            else :
                #pdb.set_trace()
                ws['G175'] = 'NOK'
                log.warning('Testing NOK for transverse z on product {}'.format(self.SN))

        except Exception:
            log.error('Transverse result failed on: {}'.format(self.SN))
        









