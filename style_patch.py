from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl.utils import range_boundaries


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)

        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        for merge_range in self._merged_cells[:]:
            merge_min_col, merge_min_row, merge_max_col, merge_max_row = range_boundaries(merge_range)
            issubset = ((min_row <= merge_min_row <= merge_max_row <= max_row) and
                        (min_col <= merge_min_col <= merge_max_col <= max_col))
            if issubset:
                # extend the existing range
                self._merged_cells.remove(merge_range)
                self._merged_cells.append(range_string)
                break
            issuperset = ((merge_min_row <= min_row <= max_row <= merge_max_row) and
                          (merge_min_col <= min_col <= max_col <= merge_max_col))
            if issuperset:
                # ignore the range (already extended)
                break
            intercept = (((merge_min_row <= min_row <= merge_max_row) or
                          (min_row <= merge_max_row <= max_row)) and
                         ((merge_min_col <= min_col <= merge_max_col) or
                          (min_col <= merge_max_col <= max_col)))
            if intercept:
                fmt = "The range {0} intercept the merged cells: {1}"
                raise ValueError(fmt.format(range_string, merge_range))
        else:
            # merge cells
            self._merged_cells.append(range_string)



        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells

patch_worksheet()